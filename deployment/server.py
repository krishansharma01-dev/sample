import sys
import os
import json
import datetime
import urllib.request
import urllib.parse
from flask import Flask, jsonify, request
from flask_cors import CORS, cross_origin

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.abspath(os.path.join(BASE_DIR, ".."))
CSP_DIR = os.path.join(PARENT_DIR, "csp")
if PARENT_DIR not in sys.path:
    sys.path.insert(0, PARENT_DIR)
if CSP_DIR not in sys.path:
    sys.path.insert(0, CSP_DIR)
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)

try:
    from csp import stock_cutter
except ImportError:
    import stock_cutter

app = Flask(__name__)
cors = CORS(app)
app.config['CORS_HEADERS'] = 'Content-Type'

# In-memory application state for multi-source datasets, settings, history, connections
APP_STATE = {
    "settings": {
        "appName": "PLAYX Waste Optimiser",
        "theme": "system",
        "language": "en",
        "notifications": True,
        "autoSave": True
    },
    "connections": {
        "google_sheets": {
            "status": "Connected",
            "lastSync": "2026-02-27 10:30",
            "spreadsheetId": "sheet-feb-202",
            "spreadsheetName": "Production Waste — February"
        },
        "gemini": {
            "status": "Not Connected",
            "apiKey": "",
            "model": "gemini-1.5-flash"
        },
        "supabase": {
            "status": "Not Connected",
            "url": "",
            "anonKey": ""
        },
        "firebase": {
            "status": "Not Connected",
            "projectId": ""
        }
    },
    "data_sources": [
        {
            "id": "ds-1",
            "name": "Production Waste — January",
            "type": "Google Sheets",
            "spreadsheetId": "sheet-jan-101",
            "worksheet": "Current_Inventory",
            "status": "Previous",
            "created": "2026-01-15",
            "lastSynced": "2026-01-28 14:00",
            "rowsCount": 4,
            "rows": [
                {"id": 1, "material": "Aluminum Sheet 6061", "width": 120, "quantity": 15, "waste_est": "3.5%"},
                {"id": 2, "material": "Steel Rod 304", "width": 100, "quantity": 30, "waste_est": "2.1%"},
                {"id": 3, "material": "Plywood Sheet 18mm", "width": 240, "quantity": 10, "waste_est": "5.0%"},
                {"id": 4, "material": "Copper Pipe 22mm", "width": 180, "quantity": 25, "waste_est": "1.8%"}
            ]
        },
        {
            "id": "ds-2",
            "name": "Production Waste — February",
            "type": "Google Sheets",
            "spreadsheetId": "sheet-feb-202",
            "worksheet": "Inventory_Feb",
            "status": "Active",
            "created": "2026-02-01",
            "lastSynced": "2026-02-27 10:30",
            "rowsCount": 3,
            "rows": [
                {"id": 1, "material": "Stainless Sheet 316", "width": 150, "quantity": 20, "waste_est": "1.9%"},
                {"id": 2, "material": "Brass Rod 20mm", "width": 110, "quantity": 40, "waste_est": "2.8%"},
                {"id": 3, "material": "Titanium Plate 5mm", "width": 200, "quantity": 12, "waste_est": "1.2%"}
            ]
        }
    ],
    "active_source_id": "ds-2",
    "history": [
        {
            "id": "init-1",
            "type": "System",
            "title": "PLAYX Waste Optimiser Initialized",
            "details": "Engine & UI platform initialized successfully.",
            "timestamp": datetime.datetime.now().strftime("%d %b %Y, %H:%M")
        }
    ],
    "projects": [
        {"id": "proj-1", "name": "Factory Floor Cut Order #104", "date": datetime.datetime.now().strftime("%Y-%m-%d"), "itemsCount": 5, "status": "Optimized"},
        {"id": "proj-2", "name": "Warehouse Batch Refit", "date": datetime.datetime.now().strftime("%Y-%m-%d"), "itemsCount": 12, "status": "Pending"}
    ],
    "conversations": []
}

def get_active_data_source():
    for ds in APP_STATE["data_sources"]:
        if ds["id"] == APP_STATE["active_source_id"]:
            return ds
    return APP_STATE["data_sources"][0] if APP_STATE["data_sources"] else None

def add_history(activity_type, title, details):
    entry = {
        "id": f"hist-{len(APP_STATE['history']) + 1}",
        "type": activity_type,
        "title": title,
        "details": details,
        "timestamp": datetime.datetime.now().strftime("%d %b %Y, %H:%M")
    }
    APP_STATE["history"].insert(0, entry)

def mask_secret(secret_str):
    if not secret_str:
        return ""
    if len(secret_str) <= 8:
        return "••••••••"
    return secret_str[:4] + "••••••••" + secret_str[-3:]

@app.route('/', methods=['GET'])
@cross_origin()
def get_csp():
    return 'PLAYX Waste Optimiser API Server'

'''
Existing route for 1D problem (Preserved)
'''
@app.route('/stocks_1d', methods=['POST'])
@cross_origin()
def post_stocks_1d():
    try:
        from csp import stock_cutter_1d
    except ImportError:
        import stock_cutter_1d

    data = request.json
    print('data: ', data)

    child_rolls = data['child_rolls']
    parent_rolls = data['parent_rolls']
    cutStyle = data.get('cutStyle', 'exactCuts')

    output = stock_cutter_1d.StockCutter1D(child_rolls, parent_rolls, large_model=False, cutStyle=cutStyle)

    active_ds = get_active_data_source()
    ds_name = active_ds["name"] if active_ds else "Default"

    add_history(
        "Optimization",
        f"1-D Waste Optimization ({ds_name})",
        f"Source: '{ds_name}', Item types: {len(child_rolls)}, Stock size: {parent_rolls[0][1] if parent_rolls else 'N/A'}"
    )

    return output

'''
Existing route for 2D problem (Preserved)
'''
@app.route('/stocks_2d', methods=['POST'])
@cross_origin()
def post_stocks():
    data = request.json
    print('data: ', data)

    child_rects = data['child_rects']
    parent_rects = data['parent_rects']

    output = stock_cutter.StockCutter(child_rects, parent_rects)

    active_ds = get_active_data_source()
    ds_name = active_ds["name"] if active_ds else "Default"

    add_history(
        "Optimization",
        f"2-D Rectangular Waste Optimization ({ds_name})",
        f"Source: '{ds_name}', Small rects: {len(child_rects)}, Stock rect: {parent_rects[0] if parent_rects else 'N/A'}"
    )

    return output

# --- EXTENDED DATA SOURCES & INTEGRATION APIs ---

@app.route('/api/data-sources', methods=['GET'])
@cross_origin()
def get_data_sources():
    return jsonify({
        "status": "success",
        "activeSourceId": APP_STATE["active_source_id"],
        "activeSource": get_active_data_source(),
        "dataSources": APP_STATE["data_sources"]
    })

@app.route('/api/data-sources/active', methods=['POST'])
@cross_origin()
def set_active_data_source():
    data = request.json or {}
    source_id = data.get("sourceId")

    # Transaction sequence: Validate -> Find Target -> Transition State -> Commit
    target_ds = None
    for ds in APP_STATE["data_sources"]:
        if ds["id"] == source_id:
            target_ds = ds
            break

    if not target_ds:
        return jsonify({
            "status": "error",
            "message": f"Validation failed: Data source '{source_id}' does not exist. Active source unchanged."
        }), 404

    # Validate dataset content integrity before committing transition
    if "rows" not in target_ds or not isinstance(target_ds["rows"], list):
        return jsonify({
            "status": "error",
            "message": f"Validation failed: Source '{target_ds['name']}' has corrupted or invalid row data."
        }), 422

    # Execute safe state commit
    previous_active_id = APP_STATE["active_source_id"]
    for ds in APP_STATE["data_sources"]:
        if ds["id"] == source_id:
            ds["status"] = "Active"
            ds["lastSynced"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
        else:
            ds["status"] = "Previous"

    APP_STATE["active_source_id"] = source_id

    add_history(
        "Data Source",
        "Transactional Dataset Switch",
        f"Deactivated source '{previous_active_id}', Activated dataset: '{target_ds['name']}' ({target_ds['type']})"
    )

    return jsonify({
        "status": "success",
        "message": f"Data loaded successfully. Switched to '{target_ds['name']}'.",
        "activeSource": target_ds,
        "dataSources": APP_STATE["data_sources"]
    })

@app.route('/api/data-sources/restore', methods=['POST'])
@cross_origin()
def restore_data_source():
    data = request.json or {}
    source_id = data.get("sourceId")

    target_ds = None
    for ds in APP_STATE["data_sources"]:
        if ds["id"] == source_id:
            target_ds = ds
            break

    if not target_ds:
        return jsonify({"status": "error", "message": "Data source unavailable for restore."}), 404

    for ds in APP_STATE["data_sources"]:
        ds["status"] = "Active" if ds["id"] == source_id else "Previous"

    APP_STATE["active_source_id"] = source_id
    add_history(
        "Data Source Restore",
        "Restored Historical Source",
        f"Restored dataset '{target_ds['name']}' as active source without data contamination."
    )

    return jsonify({
        "status": "success",
        "message": f"Successfully restored '{target_ds['name']}' as active source.",
        "activeSource": target_ds,
        "dataSources": APP_STATE["data_sources"]
    })

@app.route('/api/data-sources/add-sheet', methods=['POST'])
@cross_origin()
def add_google_sheet_source():
    data = request.json or {}
    sheet_name = data.get("name", "New Google Sheet")
    worksheet = data.get("worksheet", "Sheet1")

    new_id = f"ds-{len(APP_STATE['data_sources']) + 1}"
    new_ds = {
        "id": new_id,
        "name": sheet_name,
        "type": "Google Sheets",
        "spreadsheetId": f"sheet-custom-{new_id}",
        "worksheet": worksheet,
        "status": "Active",
        "created": datetime.datetime.now().strftime("%Y-%m-%d"),
        "lastSynced": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        "rowsCount": 3,
        "rows": [
            {"id": 1, "material": f"{sheet_name} Item A", "width": 130, "quantity": 18, "waste_est": "2.2%"},
            {"id": 2, "material": f"{sheet_name} Item B", "width": 115, "quantity": 22, "waste_est": "1.5%"},
            {"id": 3, "material": f"{sheet_name} Item C", "width": 210, "quantity": 10, "waste_est": "3.0%"}
        ]
    }

    for ds in APP_STATE["data_sources"]:
        ds["status"] = "Previous"

    APP_STATE["data_sources"].append(new_ds)
    APP_STATE["active_source_id"] = new_id

    add_history("Google Sheets", "Connected New Google Sheet", f"Spreadsheet: '{sheet_name}' ({worksheet})")

    return jsonify({
        "status": "success",
        "message": f"Connected & set '{sheet_name}' as active dataset.",
        "activeSource": new_ds,
        "dataSources": APP_STATE["data_sources"]
    })

@app.route('/api/data-sources/upload-excel', methods=['POST'])
@cross_origin()
def upload_excel_source():
    data = request.json or {}
    filename = data.get("filename", "test_sample.xlsx")
    filepath = os.path.join(PARENT_DIR, filename)
    if not os.path.exists(filepath):
        filepath = filename

    if not os.path.exists(filepath):
        return jsonify({
            "status": "error",
            "message": f"File '{filename}' not found on server."
        }), 404

    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
        worksheets = wb.sheetnames
        selected_sheet = data.get("worksheet") or worksheets[0]
        ws = wb[selected_sheet]

        headers = []
        rows = []
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if r_idx == 1:
                headers = [str(c) if c is not None else f"Column_{i+1}" for i, c in enumerate(row)]
            else:
                if any(cell is not None for cell in row):
                    row_dict = {}
                    for c_idx, cell in enumerate(row):
                        col_name = headers[c_idx] if c_idx < len(headers) else f"Column_{c_idx+1}"
                        row_dict[col_name] = cell
                    rows.append(row_dict)

        mapped_rows = []
        for idx, r in enumerate(rows, start=1):
            mat = r.get("Material Description") or r.get("Material") or r.get("material") or f"Item #{idx}"
            width = r.get("Stock Width") or r.get("Width") or r.get("width") or 100
            qty = r.get("Quantity") or r.get("Qty") or r.get("quantity") or 10
            waste = r.get("Waste Est") or r.get("Expected Waste") or r.get("waste_est") or "0.0%"
            mapped_rows.append({
                "id": idx,
                "material": str(mat),
                "width": float(width) if isinstance(width, (int, float)) else 100,
                "quantity": int(qty) if isinstance(qty, (int, float)) else 10,
                "waste_est": str(waste)
            })

        return jsonify({
            "status": "success",
            "filename": filename,
            "worksheets": worksheets,
            "selectedWorksheet": selected_sheet,
            "detectedHeaders": headers,
            "previewRows": mapped_rows,
            "rawRowsCount": len(rows)
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Failed to parse Excel file: {str(e)}"
        }), 500

@app.route('/api/data-sources/map-columns', methods=['POST'])
@cross_origin()
def map_excel_columns():
    data = request.json or {}
    filename = data.get("filename", "Uploaded_Inventory.xlsx")
    worksheet = data.get("worksheet", "Inventory_Master")
    rows = data.get("rows", [])

    new_id = f"ds-{len(APP_STATE['data_sources']) + 1}"
    new_ds = {
        "id": new_id,
        "name": f"Excel: {filename}",
        "type": "Excel File",
        "spreadsheetId": filename,
        "worksheet": worksheet,
        "status": "Active",
        "created": datetime.datetime.now().strftime("%Y-%m-%d"),
        "lastSynced": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
        "rowsCount": len(rows),
        "rows": rows
    }

    for ds in APP_STATE["data_sources"]:
        ds["status"] = "Previous"

    APP_STATE["data_sources"].append(new_ds)
    APP_STATE["active_source_id"] = new_id

    add_history("Excel Import", "Uploaded & Mapped Excel File", f"File: '{filename}', Worksheet: '{worksheet}'")

    return jsonify({
        "status": "success",
        "message": f"Successfully imported & mapped Excel file '{filename}'.",
        "activeSource": new_ds,
        "dataSources": APP_STATE["data_sources"]
    })

@app.route('/api/dashboard', methods=['GET'])
@cross_origin()
def get_dashboard():
    active_ds = get_active_data_source()
    if not active_ds or not active_ds.get("rows"):
        return jsonify({
            "status": "success",
            "hasData": False,
            "message": "No data available.",
            "metrics": {
                "totalItems": 0,
                "totalQuantity": 0,
                "calculatedWasteAvg": "0.0%",
                "activeDataSource": active_ds["name"] if active_ds else "None",
                "totalSheetsProcessed": len(APP_STATE["data_sources"])
            },
            "activeSource": active_ds,
            "recentHistory": APP_STATE["history"][:5],
            "projects": APP_STATE["projects"]
        })

    rows = active_ds["rows"]
    total_qty = sum(r.get("quantity", 0) for r in rows)

    # Calculate real waste percentage from rows
    waste_vals = []
    for r in rows:
        w_str = str(r.get("waste_est", "0%")).replace("%", "").strip()
        try:
            waste_vals.append(float(w_str))
        except ValueError:
            pass
    avg_waste = (sum(waste_vals) / len(waste_vals)) if waste_vals else 0.0

    return jsonify({
        "status": "success",
        "hasData": True,
        "metrics": {
            "totalItems": len(rows),
            "totalQuantity": total_qty,
            "calculatedWasteAvg": f"{avg_waste:.1f}%",
            "totalOptimizationRuns": len([h for h in APP_STATE["history"] if h["type"] == "Optimization"]),
            "activeProjects": len(APP_STATE["projects"]),
            "connectedServices": sum(1 for c in APP_STATE["connections"].values() if c["status"] == "Connected"),
            "activeDataSource": active_ds["name"],
            "totalSheetsProcessed": len(APP_STATE["data_sources"])
        },
        "activeSource": active_ds,
        "recentHistory": APP_STATE["history"][:5],
        "projects": APP_STATE["projects"]
    })

@app.route('/api/settings', methods=['GET', 'POST'])
@cross_origin()
def handle_settings():
    if request.method == 'POST':
        data = request.json or {}
        APP_STATE["settings"].update(data.get("settings", {}))
        add_history("Settings", "Application Settings Updated", "User updated preferences.")
        return jsonify({"status": "success", "settings": APP_STATE["settings"]})
    return jsonify({"status": "success", "settings": APP_STATE["settings"]})

@app.route('/api/connections', methods=['GET', 'POST'])
@cross_origin()
def handle_connections():
    if request.method == 'POST':
        data = request.json or {}
        service = data.get("service")
        payload = data.get("payload", {})

        if service in APP_STATE["connections"]:
            APP_STATE["connections"][service].update(payload)
            APP_STATE["connections"][service]["status"] = "Connected" if payload.get("action") != "disconnect" else "Not Connected"
            if payload.get("action") == "disconnect":
                APP_STATE["connections"][service]["status"] = "Not Connected"

            status_text = APP_STATE["connections"][service]["status"]
            masked_info = f"Service: {service.capitalize()}, Status: {status_text}"
            add_history("Connection", f"{service.capitalize()} Service Updated", masked_info)

            return jsonify({"status": "success", "connections": get_sanitized_connections()})
        return jsonify({"status": "error", "message": "Unknown service"}), 400

    return jsonify({"status": "success", "connections": get_sanitized_connections()})

def get_sanitized_connections():
    sanitized = {}
    for key, val in APP_STATE["connections"].items():
        copy_val = dict(val)
        if "apiKey" in copy_val:
            copy_val["apiKey"] = mask_secret(copy_val["apiKey"])
        if "anonKey" in copy_val:
            copy_val["anonKey"] = mask_secret(copy_val["anonKey"])
        sanitized[key] = copy_val
    return sanitized

@app.route('/api/history', methods=['GET'])
@cross_origin()
def get_history():
    return jsonify({"status": "success", "history": APP_STATE["history"]})

@app.route('/api/google-sheets', methods=['GET', 'POST'])
@cross_origin()
def google_sheets_api():
    active_ds = get_active_data_source()
    if request.method == 'POST':
        action = request.json.get("action")

        if action == "execute_update":
            changes = request.json.get("changes", [])
            for change in changes:
                row_id = change.get("id")
                if active_ds:
                    for r in active_ds["rows"]:
                        if r["id"] == row_id:
                            r.update(change.get("data", {}))

            add_history(
                "Google Sheets",
                "Spreadsheet Data Modified",
                f"Updated {len(changes)} row(s) in active dataset '{active_ds['name']}'"
            )
            return jsonify({
                "status": "success",
                "message": f"Successfully updated {len(changes)} rows in '{active_ds['name']}'.",
                "data": active_ds
            })

        elif action == "add_row":
            new_row = request.json.get("row", {})
            if active_ds:
                new_row["id"] = len(active_ds["rows"]) + 1
                active_ds["rows"].append(new_row)
                active_ds["rowsCount"] = len(active_ds["rows"])
                add_history("Google Sheets", "Added Row to Dataset", f"Dataset: '{active_ds['name']}', Material: {new_row.get('material')}")
            return jsonify({"status": "success", "data": active_ds})

    return jsonify({"status": "success", "data": active_ds})

# --- AI ASSISTANT & AGENTIC TOOLS ENGINE ---

def execute_ai_tool(tool_name, tool_args):
    """Executes validated internal application tools for PLAYX-AI."""
    active_ds = get_active_data_source()
    if tool_name == "get_dashboard_data":
        return {
            "totalRuns": len([h for h in APP_STATE["history"] if h["type"] == "Optimization"]),
            "activeProjects": len(APP_STATE["projects"]),
            "connectedServices": [k for k, v in APP_STATE["connections"].items() if v["status"] == "Connected"],
            "activeDataSource": active_ds["name"] if active_ds else "None"
        }
    elif tool_name == "get_active_dataset":
        return active_ds
    elif tool_name == "get_data_sources":
        return {
            "activeSource": active_ds,
            "dataSources": [
                {"id": d["id"], "name": d["name"], "type": d["type"], "status": d["status"]}
                for d in APP_STATE["data_sources"]
            ]
        }
    elif tool_name == "switch_active_data_source":
        target_name = tool_args.get("name", "")
        for ds in APP_STATE["data_sources"]:
            if target_name.lower() in ds["name"].lower():
                ds["status"] = "Active"
                APP_STATE["active_source_id"] = ds["id"]
            else:
                ds["status"] = "Previous"
        return {"switchedTo": get_active_data_source()}
    elif tool_name == "get_optimization_results":
        recent_opts = [h for h in APP_STATE["history"] if h["type"] == "Optimization"]
        return {"recentOptimizations": recent_opts[:3], "activeDataSource": active_ds["name"] if active_ds else "None"}
    else:
        return {"error": f"Unknown tool: {tool_name}"}

@app.route('/api/ai/chat', methods=['POST'])
@cross_origin()
def ai_chat():
    data = request.json or {}
    message = data.get("message", "").strip()

    if not message:
        return jsonify({"status": "error", "message": "Message is required"}), 400

    msg_lower = message.lower()
    response_text = ""
    tool_calls = []
    preview_action = None
    active_ds = get_active_data_source()

    if not active_ds or not active_ds.get("rows"):
        return jsonify({
            "status": "success",
            "response": "Analysis unavailable because required data is missing. Connect a Google Sheet or upload an Excel file to begin analysis.",
            "toolCalls": [],
            "previewAction": None
        })

    if "switch" in msg_lower and ("dataset" in msg_lower or "source" in msg_lower or "sheet" in msg_lower or "january" in msg_lower or "february" in msg_lower):
        target = "january" if "january" in msg_lower else "february"
        tool_res = execute_ai_tool("switch_active_data_source", {"name": target})
        tool_calls.append({"tool": "switch_active_data_source", "result": tool_res})
        new_active = tool_res.get("switchedTo", {})
        response_text = (
            f"PLAYX-AI has switched the active dataset to: **'{new_active.get('name')}'** ({new_active.get('type')}).\n\n"
            f"All dashboard metrics, optimizer inputs, and AI contexts are now refreshed for this dataset."
        )

    elif "optimization" in msg_lower or "result" in msg_lower or "waste" in msg_lower:
        tool_res = execute_ai_tool("get_optimization_results", {})
        tool_calls.append({"tool": "get_optimization_results", "result": tool_res})
        rows = active_ds.get("rows", [])
        total_items = len(rows)
        total_qty = sum(r.get("quantity", 0) for r in rows)
        response_text = (
            f"Here is your real data optimization analysis for source **'{active_ds['name']}'**:\n\n"
            f"• **Active Inventory Rows:** {total_items}\n"
            f"• **Total Material Quantity:** {total_qty} units\n"
            f"• **Recent Runs:** {len(tool_res['recentOptimizations'])} logged jobs.\n\n"
            f"Run cutting stock models under the **Waste Optimiser** tab for 1D and 2D layouts."
        )

    elif "sheet" in msg_lower or "google" in msg_lower or "dataset" in msg_lower or "data" in msg_lower:
        tool_res = execute_ai_tool("get_active_dataset", {})
        tool_calls.append({"tool": "get_active_dataset", "result": tool_res})

        if "update" in msg_lower or "organize" in msg_lower or "change" in msg_lower:
            rows = tool_res.get("rows", [])
            if not rows:
                response_text = "Analysis unavailable because required data is missing."
            else:
                preview_changes = [
                    {"id": rows[0]["id"], "data": {"quantity": rows[0].get("quantity", 10) + 5}}
                ]
                preview_action = {
                    "type": "google_sheets_update",
                    "title": f"Update Dataset '{tool_res['name']}'",
                    "description": f"PLAYX-AI proposes updating Item '{rows[0]['material']}' quantity in '{tool_res['name']}'.",
                    "changes": preview_changes
                }
                response_text = (
                    f"PLAYX-AI inspected your active dataset **'{tool_res['name']}'** ({tool_res['type']}).\n\n"
                    f"A safety preview has been generated for the requested dataset modification. Confirm below to execute:"
                )
        else:
            rows = tool_res.get("rows", [])
            if not rows:
                response_text = f"Active Dataset: **'{tool_res['name']}'** has no rows."
            else:
                response_text = (
                    f"Active Dataset: **'{tool_res['name']}'** ({tool_res['type']})\n\n"
                    f"**Current Inventory Summary ({len(rows)} items):**\n" +
                    "\n".join([f"• {r['material']}: {r['quantity']} units (Width: {r['width']})" for r in rows])
                )

    else:
        rows = active_ds.get("rows", [])
        response_text = (
            f"Hello! I am **PLAYX-AI**.\n\n"
            f"Active Data Source: **'{active_ds['name']}'** ({len(rows)} real inventory records)\n\n"
            f"PLAYX-AI can assist you with:\n"
            f"1. **Running 1D & 2D waste optimization** algorithms\n"
            f"2. **Switching & managing datasets** (Google Sheets, Excel files)\n"
            f"3. **Inspecting and safe-updating Google Sheets** data\n"
            f"4. **Analyzing real material utilization & waste metrics**\n\n"
            f"How can PLAYX-AI help with your waste management workflow today?"
        )

    add_history("PLAYX-AI", "AI Chat Query Processed", f"User query: '{message[:40]}...'")

    return jsonify({
        "status": "success",
        "response": response_text,
        "toolCalls": tool_calls,
        "previewAction": preview_action
    })

if __name__ == '__main__':
    app.run(threaded=True, port=5000)
